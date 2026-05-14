"""
================================================================================
BUILD_EXCEL_EXPLAINED.py — Part 1 of 3
================================================================================
Builds the range-accrual pricing workbook. Read this top-to-bottom.

Three files together make the complete walkthrough:

    Part 1 (this file):  imports, helpers, market data dicts
    Part 2:              SABR formula builders, sheet-building functions
    Part 3:              main entry point that ties everything together

If you only want to RUN the model, use build.py — this trio is for reading.

================================================================================
HIGH-LEVEL ARCHITECTURE
================================================================================
The script does five things, in order:

1.  IMPORT LIBRARIES — bring in pre-written code we need
    (openpyxl writes .xlsx files; datetime handles dates; etc.)

2.  DEFINE HELPER FUNCTIONS — small reusable bits of logic:
      - date math (year fractions, third-Wednesday-of-month)
      - Excel cell-writing wrappers (styling, fonts, fills)
      - SABR-vol formula builders (return strings of Excel formulas)

3.  COLLECT MARKET DATA — Python dicts (key→value lookups) hold the
    inputs for one valuation date: SOFR rate, FOMC step quotes,
    SOFR futures prices, par swap rates, ATM vol surface, SABR params,
    plus trade terms (notional, coupon, barrier K, dates).

4.  BUILD SHEETS — for each sheet (MarketData, YieldCurve, DailyDFs,
    DailyFixings, PeriodSummary, Valuation, README), one Python function
    writes the data and formulas. They write Excel FORMULAS, not values
    — Excel does the actual arithmetic when the workbook opens.

5.  SAVE THE FILE — openpyxl serialises everything to disk as .xlsx.

================================================================================
PYTHON BASICS YOU WILL SEE
================================================================================
  *  def starts a function definition. The body is INDENTED.
       def my_func(x):
           return x + 1
  *  Strings can use single or double quotes. f"...{x}..." is an
     "f-string": Python substitutes the value of x where {x} appears.
  *  Lists use square brackets: [1, 2, 3]
  *  Dicts use curly braces: {'name': 'Alice', 'age': 30}
  *  Tuples use parentheses: (1, 2). Immutable list, basically.
  *  for x in collection:  loops over each item.
  *  if / elif / else for conditionals.
  *  Indentation IS the syntax. No braces.
  *  # starts a comment to end of line.
  *  Triple-quoted strings are docstrings (function descriptions).

================================================================================
THE BIG IDEA
================================================================================
Every cell in the workbook is one of:
   (a) a hardcoded INPUT  (market data — coloured blue)
   (b) an Excel FORMULA  referring to other cells (black, or green for
       cross-sheet references)
This script writes (a) directly and writes (b) as STRINGS starting with
"=...". When Excel/LibreOffice opens the file, it evaluates the formulas.
================================================================================
"""

# ═══════════════════════════════════════════════════════════════════════════
# PART 1.1 — IMPORTS
# ═══════════════════════════════════════════════════════════════════════════
# "import X" makes X.foo() callable.
# "from X import Y" pulls Y directly into our namespace.

from openpyxl import Workbook                  # main Excel writer
from openpyxl.styles import (                  # cell styling
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter   # 1 -> "A", 27 -> "AA"
from datetime import date, timedelta           # date arithmetic
import sys                                     # command-line args
import os                                      # filesystem ops


# ═══════════════════════════════════════════════════════════════════════════
# PART 1.2 — STYLE OBJECTS (CONSTANTS)
# ═══════════════════════════════════════════════════════════════════════════
# These are openpyxl Font/Fill/Alignment "specs" — created once, used
# everywhere. UPPER_CASE names are a Python convention for constants.
# Colours are hex RGB strings. ARGB if 8 chars (alpha first).

INPUT_FONT    = Font(name='Calibri', size=11, color='0000FF')         # blue
FORMULA_FONT  = Font(name='Calibri', size=11, color='000000')         # black
HEADER_FONT   = Font(name='Calibri', size=11, color='000000', bold=True)
SUBTITLE_FONT = Font(name='Calibri', size=12, color='000000', bold=True)
TITLE_FONT    = Font(name='Calibri', size=14, color='000000', bold=True)
CROSS_FONT    = Font(name='Calibri', size=11, color='006400')         # green
INPUT_FILL    = PatternFill('solid', fgColor='FFF2CC')                # pale yellow
SECTION_FILL  = PatternFill('solid', fgColor='D9E1F2')                # pale blue
CENTER_ALIGN  = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT_ALIGN   = Alignment(horizontal='right',  vertical='center')


# ═══════════════════════════════════════════════════════════════════════════
# PART 1.3 — TINY HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def setcell(ws, row, col, value, font=None, *, fmt=None,
            fill=None, align=None, border=None):
    """
    Write one styled Excel cell. `ws` is an openpyxl worksheet.

    The lone `*` in the signature means parameters AFTER it must be
    passed by NAME, not position. So you write:
        setcell(ws, 5, 2, 0.04, fmt='0.000%')
    not
        setcell(ws, 5, 2, 0.04, '0.000%')   # would be a TypeError

    `value` may be a number, string, date, or — crucially — a string
    starting with "=", which Excel interprets as a formula.
    """
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font          = font
    if fmt:    c.number_format = fmt
    if fill:   c.fill          = fill
    if align:  c.alignment     = align
    if border: c.border        = border
    return c


def add_years(d, n):
    """
    Add n calendar years to a date.

    `d.replace(year=...)` returns a NEW date with the year changed
    (dates in Python are immutable). It raises ValueError if we ask
    for "Feb 29, 2027" — there is no such date. The try/except
    catches that and falls back to Feb 28.
    """
    try:
        return d.replace(year=d.year + n)
    except ValueError:
        return d.replace(year=d.year + n, month=2, day=28)


def third_wednesday(year, month):
    """
    Third Wednesday of a given (year, month). Used for SOFR-futures
    expiry dates (IMM dates).

    `weekday()` returns 0=Monday, ..., 6=Sunday. So Wednesday is 2.
    `(2 - first.weekday()) % 7` is the day-offset from the 1st to
    the first Wednesday. Adding 14 jumps two more weeks to land on
    the THIRD Wednesday.
    """
    first = date(year, month, 1)
    return first + timedelta(days=(2 - first.weekday()) % 7 + 14)


def daycount_30_360(d1, d2):
    """30/360 day-count convention (used by the coupon leg)."""
    d1d, d1m, d1y = d1.day, d1.month, d1.year
    d2d, d2m, d2y = d2.day, d2.month, d2.year
    d1d = min(d1d, 30)
    if d1d == 30:
        d2d = min(d2d, 30)
    return ((d2y - d1y) * 360 + (d2m - d1m) * 30 + (d2d - d1d)) / 360.0


def daycount_act_360(d1, d2):
    """ACT/360 day-count convention (used by SOFR/curve)."""
    return (d2 - d1).days / 360.0


# ═══════════════════════════════════════════════════════════════════════════
# PART 1.4 — MARKET DATA
# ═══════════════════════════════════════════════════════════════════════════
# Each valuation date is captured as a "dict" — Python's key→value map.
# We return the dict from a function so the same builder can use either
# date by calling make_data_may7() or make_data_apr21().
#
# Trade terms are SHARED across dates (same trade, just different val date).
# Below we define them once and merge into each dict.

TRADE_TERMS = {
    'notional':     1_370_000,                    # USD notional
    'coupon':       0.0532,                       # 5.32%
    'K':            0.042,                        # 4.2% upper barrier
    'leg3_start':   date(2026, 10, 24),           # 1st accrual period start
    'leg3_end':     date(2031, 4, 24),            # final maturity
    'beta':         0.40,                         # SABR β  (rates convention)
    'delta_K':      0.0001,                       # put-spread half-width (1bp)
}

# Underscores in 1_370_000 are just digit-group separators — Python ignores
# them. Improves readability for big numbers.


def make_data_may7():
    """
    Build the market-data dict for valuation = 7 May 2026.
    Each list contains tuples (immutable groups of values). Inside the
    dict we have multiple categories:
       - sofr_1d:    overnight SOFR rate
       - step_quotes: FOMC step quotes (intervals + rate)
       - futures:    SOFR 3M futures with conv-adj
       - swaps:      par swap rates by tenor
       - atm_10Y:    ATM vol surface for 10Y CMS tail
       - sabr_10Y:   ρ and ν parameters by expiry
    """
    d = {
        'val_date': date(2026, 5, 7),
        'sofr_1d':  0.036313,

        # FOMC step quotes: each tuple is (start, end, rate). These are
        # piecewise-constant forward rates between FOMC meeting dates.
        'step_quotes': [
            (date(2026, 6, 18),  date(2026, 7, 30),  0.0364),
            (date(2026, 7, 30),  date(2026, 9, 17),  0.0362),
            (date(2026, 9, 17),  date(2026, 10, 29), 0.0361),
            (date(2026, 10, 29), date(2026, 12, 10), 0.0363),
            (date(2026, 12, 10), date(2027, 1, 28),  0.0366),
            (date(2027, 1, 28),  date(2027, 3, 18),  0.0368),
            (date(2027, 3, 18),  date(2027, 4, 29),  0.0369),
            (date(2027, 4, 29),  date(2027, 6, 10),  0.0370),
        ],

        # SOFR 3-month futures: (year, IMM month, price, convexity adj %)
        # Forward rate = (100 - price)/100 - conv_adj/100
        'futures': [
            (2027, 6,  96.32,    0.0048),
            (2027, 9,  96.3794,  0.0076),
            (2027, 12, 96.44,    0.0103),
            (2028, 3,  96.4651,  0.0135),
            (2028, 6,  96.4613,  0.0174),
            (2028, 9,  96.4445,  0.0217),
            (2028, 12, 96.4167,  0.0269),
            (2029, 3,  96.385,   0.033),
        ],

        # Par swap rates: tenor (years) -> rate (decimal)
        'swaps': {
            4:0.03659, 5:0.036867, 6:0.037296, 7:0.03776, 8:0.038215,
            9:0.038669, 10:0.039119, 11:0.039564, 12:0.039995,
            15:0.041081, 20:0.041976, 25:0.04206, 30:0.041764,
        },

        # ATM volatility (10Y CMS tail), indexed by option expiry in years
        'atm_10Y': [
            (1/12,  0.107076), (3/12, 0.120642), (6/12, 0.126932),
            (9/12,  0.128887), (1.0,  0.130858), (2.0,  0.132248),
            (3.0,   0.131838), (4.0,  0.130748), (5.0,  0.129439),
            (7.0,   0.126310), (10.0, 0.123830), (15.0, 0.120768),
            (20.0,  0.125042), (25.0, 0.122842), (30.0, 0.134111),
        ],

        # SABR (ρ, ν) per expiry. β is set in TRADE_TERMS at 0.40.
        'sabr_10Y': [
            (1/12, 0.0714, 2.8362), (3/12, 0.0601, 1.9507),
            (6/12, 0.0575, 1.1387), (1.0,  0.0491, 0.7741),
            (2.0,  0.062,  0.4695), (5.0,  0.163,  0.3135),
            (7.0,  0.205,  0.291),  (10.0, 0.323,  0.245),
            (20.0, 0.322,  0.219),  (30.0, 0.282,  0.215),
        ],
    }
    # ** unpacks one dict into another. So `**TRADE_TERMS` adds all of its
    # entries into the returned dict.
    return {**d, **TRADE_TERMS}


def make_data_apr21():
    """Same shape as may7, but with the Apr 21 market snapshot."""
    d = {
        'val_date': date(2026, 4, 21),
        'sofr_1d':  0.036386,
        'step_quotes': [
            (date(2026, 4, 30),  date(2026, 6, 18),  0.0365),
            (date(2026, 6, 18),  date(2026, 7, 30),  0.0366),
            (date(2026, 7, 30),  date(2026, 9, 17),  0.0365),
            (date(2026, 9, 17),  date(2026, 10, 29), 0.0361),
            (date(2026, 10, 29), date(2026, 12, 10), 0.0359),
            (date(2026, 12, 10), date(2027, 1, 28),  0.0356),
            (date(2027, 1, 28),  date(2027, 3, 18),  0.0353),
        ],
        'futures': [
            (2027, 3,  96.48,    0.0027),
            (2027, 6,  96.5329,  0.0048),
            (2027, 9,  96.5992,  0.0076),
            (2027, 12, 96.6431,  0.0103),
            (2028, 3,  96.6413,  0.0135),
            (2028, 6,  96.6151,  0.0174),
            (2028, 9,  96.5822,  0.0217),
            (2028, 12, 96.5455,  0.0269),
            (2029, 3,  96.5095,  0.033),
        ],
        'swaps': {
            4:0.035309, 5:0.035659, 6:0.036155, 7:0.03669,
            8:0.037217, 9:0.037732, 10:0.038235, 11:0.038725,
            12:0.039195, 15:0.040374, 20:0.041392, 25:0.041542,
            30:0.041279,
        },
        'atm_10Y': [
            (1/12,  0.10628), (3/12, 0.10956), (6/12, 0.11536),
            (9/12,  0.11767), (1.0,  0.11995), (2.0,  0.12167),
            (3.0,   0.12132), (4.0,  0.12006), (5.0,  0.11891),
            (7.0,   0.11681), (10.0, 0.11412), (15.0, 0.11324),
            (20.0,  0.11690), (30.0, 0.12602),
        ],
        'sabr_10Y': [
            (1/12, 0.0714, 3.1362), (3/12, 0.0601, 2.1007),
            (6/12, 0.0575, 1.1887), (1.0,  0.0491, 0.7741),
            (2.0,  0.062,  0.4695), (5.0,  0.163,  0.3135),
            (7.0,  0.205,  0.291),  (10.0, 0.323,  0.245),
            (20.0, 0.322,  0.219),  (30.0, 0.282,  0.215),
        ],
    }
    return {**d, **TRADE_TERMS}
