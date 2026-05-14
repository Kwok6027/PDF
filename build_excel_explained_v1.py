"""
===============================================================================
RANGE ACCRUAL NOTE — EXCEL BUILDER (BEGINNER-FRIENDLY VERSION)
===============================================================================

This Python script creates two Excel workbooks that price a callable range
accrual note. The Excel formulas (not the Python code itself) do the math —
this script just writes the right formulas into the right cells.

WHAT YOU NEED TO RUN THIS:
  1. Python 3.8 or later     (check: open a terminal, type `python3 --version`)
  2. openpyxl library         (install: `pip install openpyxl`)
  3. LibreOffice (optional)   (only needed to "recalculate" formulas after
                               saving — Excel and LibreOffice do this on open)

HOW TO RUN:
    python3 build_excel_explained.py both        # builds both workbooks
    python3 build_excel_explained.py apr21       # builds only Apr 21
    python3 build_excel_explained.py may7        # builds only May 7

===============================================================================
PYTHON CONCEPTS YOU'LL ENCOUNTER (super-quick primer)
===============================================================================

  IMPORTS:
      `import math`                  brings in math library (math.sqrt, etc.)
      `from datetime import date`    brings in just the `date` thing

  VARIABLES:
      `x = 5`                        assigns 5 to x
      `name = "hello"`               text is wrapped in quotes (a "string")

  LISTS (ordered collections):
      my_list = [1, 2, 3]            access with my_list[0] which gives 1
                                     (Python is 0-indexed: first item is [0])
  TUPLES (immutable ordered collections):
      my_tuple = (1, 2, 3)           similar to lists but can't be changed
      Used here for (date, rate) pairs.

  DICTIONARIES (key-value storage, like a phone book):
      data = {'name': 'Alice', 'age': 30}
      data['name']                   gives 'Alice'

  FUNCTIONS (reusable code blocks):
      def add(a, b):                 defines a function called `add`
          return a + b               that takes two inputs and gives back their sum
      result = add(2, 3)             result is now 5

  F-STRINGS (text templates):
      x = 5
      msg = f"x is {x}"              msg becomes "x is 5"
                                     anything inside {} gets evaluated as Python

  LOOPS:
      for i in range(10):            i takes values 0, 1, 2, ..., 9
          print(i)
      for item in my_list:           iterate over each item in my_list
          do_something(item)

  CONDITIONALS:
      if x > 0:                      runs first block if x is positive
          ...
      elif x == 0:                   else-if x is exactly zero
          ...
      else:                          otherwise
          ...

  COMMENTS:
      # everything after the # on a line is ignored

===============================================================================
WHAT THIS SCRIPT BUILDS
===============================================================================

For each valuation date (May 7 2026 and Apr 21 2026) we build a workbook
with 7 sheets:

   1. README        — overview of the methodology
   2. MarketData    — the ONLY sheet with hardcoded numbers (rates, vols, ...)
   3. YieldCurve    — builds discount factors from market data
   4. DailyDFs      — pre-computed discount factor at every calendar day
   5. DailyFixings  — for each day in the schedule: forward 10Y CMS rate,
                      SABR-implied vol at the 4.2% barrier, digital probability
   6. PeriodSummary — aggregates daily probabilities into 18 quarterly coupons
   7. Valuation     — step-by-step walkthrough with explanations and references

The key principle: only MarketData has hardcoded values. Every other sheet
uses Excel formulas that derive everything from MarketData. So if you change
a swap rate on MarketData, the final PV updates automatically.

===============================================================================
"""

# =============================================================================
# SECTION 1 — IMPORTS
# =============================================================================
# An "import" statement brings external code into our script so we can use it.
# Without these we'd have to write everything from scratch.

import math
# `math` gives us math.sqrt, math.log, math.exp, etc.
# We barely use it here because Excel does most of the math, but a couple of
# places need it (e.g., third_wed below uses % for modular arithmetic).

from datetime import date, timedelta
# `datetime` is Python's date/time module. We import two things:
#   - `date`      represents a calendar date, e.g. date(2026, 5, 7)
#   - `timedelta` represents an interval, e.g. timedelta(days=42)
# These are how we do date arithmetic in Python.

from openpyxl import Workbook
# `openpyxl` is a third-party library for reading and writing .xlsx files.
# `Workbook` is the class that represents an entire Excel file.

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# These are formatting tools:
#   - Font:        text color, size, bold, italic
#   - PatternFill: cell background color
#   - Alignment:   left/right/center, wrap text
#   - Border/Side: cell borders (we use a thin border in some places)

from openpyxl.utils import get_column_letter
# Converts column NUMBER to LETTER. Excel columns go A, B, C, ..., Z, AA, AB...
# get_column_letter(1) returns 'A', get_column_letter(27) returns 'AA', etc.
# Useful when we want to generate cell addresses like "AG552" programmatically.


# =============================================================================
# SECTION 2 — STYLE CONSTANTS
# =============================================================================
# These are reusable styling objects we'll apply to many cells. Defining them
# once at the top keeps the code clean.
#
# Industry convention for financial spreadsheets:
#   Blue text   = hardcoded input (user might change for scenarios)
#   Black text  = formula
#   Green text  = link to another sheet within the same workbook
#   Red text    = link to an external file (we don't use this)

INPUT_FONT = Font(name='Arial', color='0000FF', size=10)     # Blue = input
FORMULA_FONT = Font(name='Arial', color='000000', size=10)   # Black = formula
XREF_FONT = Font(name='Arial', color='008000', size=10)      # Green = cross-sheet
HEADER_FONT = Font(name='Arial', bold=True, size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='C00000')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='C00000')

# PatternFill is used for cell background colors. fgColor is a 6-character
# hex code (e.g. 'FFF2CC' is light yellow). You can find hex codes online.
SECTION_FILL = PatternFill('solid', fgColor='D9D9D9')   # gray for section headers
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')     # light yellow for inputs
RESULT_FILL = PatternFill('solid', fgColor='C6E0B4')    # light green for results

# Alignment objects — how text is positioned inside a cell.
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


# =============================================================================
# SECTION 3 — UTILITY FUNCTIONS
# =============================================================================
# Small helpers used throughout the script.

def third_wed(year, month):
    """
    Return the 3rd Wednesday of a given year and month.

    SOFR 3-month futures expire on the 3rd Wednesday of the contract month
    (March, June, September, December). We need this date to convert the
    contract month (e.g., "Jun 2027") into an actual calendar date.

    HOW IT WORKS:
      - `date(year, month, 1)` gives the 1st of the month
      - `.weekday()` returns 0 for Monday, 1 for Tuesday, ..., 6 for Sunday
        so Wednesday is 2
      - We want to advance from the 1st to the FIRST Wednesday of the month,
        then add 14 more days (2 weeks) to reach the THIRD Wednesday
      - `(2 - d.weekday()) % 7` is the number of days to advance to reach
        the first Wednesday. The `% 7` (modulo) handles the "wrap around"
        when the 1st is already past Wednesday.

    Example: third_wed(2027, 6) returns date(2027, 6, 16)
    """
    d = date(year, month, 1)
    days_to_first_wed = (2 - d.weekday()) % 7
    return d + timedelta(days=days_to_first_wed + 14)


def add_months(d, m):
    """
    Add `m` months to date `d` and return the resulting date.

    Excel has EDATE() that does this, but Python's `datetime` doesn't ship with
    a built-in equivalent, so we write one. We use this for computing the
    quarterly period end dates in the leg 3 schedule.

    Edge case: if adding months gives a non-existent day (e.g. adding 1 month
    to Jan 31 would give Feb 31 which doesn't exist), we fall back to day 28.

    Example: add_months(date(2026, 10, 24), 3) returns date(2027, 1, 24)
    """
    # Integer math to figure out the new year and month
    y = d.year + (d.month - 1 + m) // 12          # // is integer division
    mo = (d.month - 1 + m) % 12 + 1               # % is the remainder
    try:
        return d.replace(year=y, month=mo)
    except ValueError:
        return d.replace(year=y, month=mo, day=28)


def setcell(ws, row, col, value, font=None, fill=None, fmt=None, align=None):
    """
    Write a value (or formula) into an Excel cell with optional styling.

    PARAMETERS:
      ws       — the worksheet object (an openpyxl sheet)
      row, col — 1-indexed cell position. row=1, col=1 is cell A1.
      value    — the content. Can be a number, text, date, or a formula
                 (formulas are STRINGS that start with '=')
      font     — Font object (optional)
      fill     — PatternFill object (optional)
      fmt      — number format string (e.g. '0.00%' for percent)
      align    — Alignment object (optional)

    The `None` default value for the optional parameters means "if not provided,
    leave it unset" so we don't have to specify every style for every cell.

    Returns the cell object so further customisation is possible.
    """
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    return c


# =============================================================================
# SECTION 4 — MARKET DATA
# =============================================================================
# These two functions return dictionaries holding all the inputs for each
# valuation date. A dictionary is like a phone book: you look up things by
# name. Example: `data['val_date']` gives back the valuation date.
#
# Why use a dictionary? Because we have many related pieces of data and we
# want to pass them all around as a single object. Without a dict we'd need
# 20 separate variables for each valuation date and the code would get messy.
#
# Note: I write floating-point rates as decimals (3.6313% = 0.036313).

def make_data_apr21():
    """Return a dictionary of market data for the April 21, 2026 valuation."""

    # `date(year, month, day)` constructor — straightforward.
    val_date = date(2026, 4, 21)

    # FOMC step quotes: each row is (start_date, end_date, rate).
    # Between Fed meetings, SOFR is assumed to be constant at the quoted rate.
    # This is a LIST of TUPLES — a list (square brackets) holding triples (parens).
    step_quotes = [
        (date(2026, 4, 30), date(2026, 6, 18), 0.0365),
        (date(2026, 6, 18), date(2026, 7, 30), 0.0366),
        (date(2026, 7, 30), date(2026, 9, 17), 0.0365),
        (date(2026, 9, 17), date(2026, 10, 29), 0.0361),
        (date(2026, 10, 29), date(2026, 12, 10), 0.0359),
        (date(2026, 12, 10), date(2027, 1, 28), 0.0356),
        (date(2027, 1, 28), date(2027, 3, 18), 0.0353),
    ]

    # SOFR 3-month futures: each row is (year, month, price, convexity_adj).
    # The "year, month" identifies the contract (e.g. 2027/Mar contract starts
    # on the 3rd Wed of Mar 2027 and ends on the 3rd Wed of June 2027).
    # The forward rate implied by the future is:
    #     forward = (100 - price)/100 - convexity_adj/100
    # The convexity adjustment corrects for the small difference between
    # a future's payoff and the underlying forward rate.
    futures_raw = [
        (2027, 3,  96.4800, 0.0027),
        (2027, 6,  96.5329, 0.0048),
        (2027, 9,  96.5992, 0.0076),
        (2027, 12, 96.6431, 0.0103),
        (2028, 3,  96.6413, 0.0135),
        (2028, 6,  96.6151, 0.0174),
        (2028, 9,  96.5822, 0.0217),
        (2028, 12, 96.5455, 0.0269),
        (2029, 3,  96.5095, 0.0330),
    ]

    # Par swap rates: each row is (tenor_in_years, rate).
    # Provided only at certain tenors (4, 5, ..., 12, 15, 20, 25, 30).
    # We'll interpolate linearly for the missing ones (13, 14, 16-19, etc.).
    swap_rates = [
        (4, 0.035309), (5, 0.035659), (6, 0.036155), (7, 0.036690),
        (8, 0.037217), (9, 0.037732), (10, 0.038235), (11, 0.038725),
        (12, 0.039195), (15, 0.040374), (20, 0.041392), (25, 0.041542),
        (30, 0.041279),
    ]

    # ATM swaption volatilities for the 10Y tail at each expiry.
    # Each row is (expiry_in_years, volatility).
    atm_10Y = [
        (1/12, 0.1062775498), (3/12, 0.1095604945), (6/12, 0.1153620649),
        (9/12, 0.1176724946), (1.0,  0.1199508224), (2.0,  0.1216738595),
        (3.0,  0.1213169981), (4.0,  0.1200644572), (5.0,  0.1189123336),
        (7.0,  0.1168146333), (10.0, 0.1141234575), (15.0, 0.1132386810),
        (20.0, 0.1169034513), (30.0, 0.1260176360),
    ]

    # SABR parameters rho (correlation) and nu (vol-of-vol) for 10Y tail.
    # Each row is (expiry_in_years, rho, nu).
    sabr_10Y = [
        (1/12, 0.0714, 3.1362), (3/12, 0.0601, 2.1007),
        (6/12, 0.0575, 1.1887), (1.0,  0.0491, 0.7741),
        (2.0,  0.0620, 0.4695), (5.0,  0.1630, 0.3135),
        (7.0,  0.2050, 0.2910), (10.0, 0.3230, 0.2450),
        (20.0, 0.3220, 0.2190), (30.0, 0.2820, 0.2150),
    ]

    # `return` sends back a dictionary (curly braces) with all our data.
    # `dict(key=value, ...)` is shorthand for {'key': value, ...}.
    return dict(
        label='Apr 21, 2026',           # for the README header
        val_date=val_date,
        sofr_1d=0.036386,                # 1-day SOFR rate
        step_quotes=step_quotes,
        futures_raw=futures_raw,
        swap_rates=swap_rates,
        atm_10Y=atm_10Y,
        sabr_10Y=sabr_10Y,
        beta=0.40,                       # SABR beta (held fixed)
        notional=1_370_000,              # USD 1.37 million (underscore is for readability)
        coupon=0.0532,                   # 5.32% annual coupon
        K=0.042,                         # 4.2% upper barrier
        leg3_start=date(2026, 10, 24),   # first day of accrual leg
        leg3_end=date(2031, 4, 24),      # last day of accrual leg
        expected_pv=181852,              # for sanity check
    )


def make_data_may7():
    """Return a dictionary of market data for the May 7, 2026 valuation."""
    val_date = date(2026, 5, 7)
    step_quotes = [
        (date(2026, 6, 18), date(2026, 7, 30), 0.0364),
        (date(2026, 7, 30), date(2026, 9, 17), 0.0362),
        (date(2026, 9, 17), date(2026, 10, 29), 0.0361),
        (date(2026, 10, 29), date(2026, 12, 10), 0.0363),
        (date(2026, 12, 10), date(2027, 1, 28), 0.0366),
        (date(2027, 1, 28), date(2027, 3, 18), 0.0368),
        (date(2027, 3, 18), date(2027, 4, 29), 0.0369),
        (date(2027, 4, 29), date(2027, 6, 10), 0.0370),
    ]
    futures_raw = [
        (2027, 6,  96.3200, 0.0048),
        (2027, 9,  96.3794, 0.0076),
        (2027, 12, 96.4400, 0.0103),
        (2028, 3,  96.4651, 0.0135),
        (2028, 6,  96.4613, 0.0174),
        (2028, 9,  96.4445, 0.0217),
        (2028, 12, 96.4167, 0.0269),
        (2029, 3,  96.3850, 0.0330),
    ]
    swap_rates = [
        (4, 0.036590), (5, 0.036867), (6, 0.037296), (7, 0.037760),
        (8, 0.038215), (9, 0.038669), (10, 0.039119), (11, 0.039564),
        (12, 0.039995), (15, 0.041081), (20, 0.041976), (25, 0.042060),
        (30, 0.041764),
    ]
    atm_10Y = [
        (1/12, 0.107076), (3/12, 0.120642), (6/12, 0.126932),
        (9/12, 0.128887), (1.0, 0.130858), (2.0, 0.132248),
        (3.0, 0.131838), (4.0, 0.130748), (5.0, 0.129439),
        (7.0, 0.126310), (10.0, 0.123830), (15.0, 0.120768),
        (20.0, 0.125042), (25.0, 0.122842), (30.0, 0.134111),
    ]
    sabr_10Y = [
        (1/12, 0.0714, 2.8362), (3/12, 0.0601, 1.9507),
        (6/12, 0.0575, 1.1387), (1.0,  0.0491, 0.7741),
        (2.0,  0.0620, 0.4695), (5.0,  0.1630, 0.3135),
        (7.0,  0.2050, 0.2910), (10.0, 0.3230, 0.2450),
        (20.0, 0.3220, 0.2190), (30.0, 0.2820, 0.2150),
    ]
    return dict(
        label='May 7, 2026', val_date=val_date, sofr_1d=0.036313,
        step_quotes=step_quotes, futures_raw=futures_raw, swap_rates=swap_rates,
        atm_10Y=atm_10Y, sabr_10Y=sabr_10Y, beta=0.40,
        notional=1_370_000, coupon=0.0532, K=0.042,
        leg3_start=date(2026, 10, 24), leg3_end=date(2031, 4, 24),
        expected_pv=172776,
    )


# =============================================================================
# SECTION 5 — BUILD THE README SHEET
# =============================================================================
# This is a "documentation" sheet that explains the workbook. No formulas,
# just text. Pretty straightforward Python: a list of lines and a loop that
# writes them.

def build_readme(wb, data):
    """
    Build the README sheet at position 0 (so it's the first one users see).

    PARAMETERS:
      wb   — the Workbook object
      data — the market data dict (we use the 'label' and 'expected_pv' from it)
    """
    # `create_sheet(name, index)` creates a new sheet at a given position.
    # index=0 means "make it the FIRST sheet".
    ws = wb.create_sheet('README', 0)

    # Set the width of column A. Excel column widths are in "character" units.
    ws.column_dimensions['A'].width = 100

    # Write the title in cell A1.
    setcell(ws, 1, 1, f"Range Accrual Note — Leg 3 Pricing — Valuation as of {data['label']}",
            TITLE_FONT)

    # A list of strings — each will go in its own row. Easier than calling
    # setcell 30 times with hardcoded row numbers.
    rows = [
        "",
        "PURPOSE",
        f"  Reproduce Leg 3 (range accrual) PV of approximately ${data['expected_pv']:,} using only market data.",
        "",
        "STRUCTURE",
        "  1. MarketData     — All hardcoded inputs (only sheet with hardcodes)",
        "  2. YieldCurve     — Build discount factor curve (DFs) from market data",
        "  3. DailyDFs       — Pre-computed DF at every calendar day (for fast lookup)",
        "  4. DailyFixings   — One row per calendar day in Leg 3; compute forward 10Y CMS,",
        "                       calibrate SABR alpha, compute digital probability",
        "  5. PeriodSummary  — Aggregate daily probabilities into 18 quarterly coupons",
        "  6. Valuation      — Step-by-step walkthrough with explanations",
        "",
        "COLOR CODE",
        "  Blue  = hardcoded input (MarketData only)",
        "  Black = formula calculation",
        "  Green = cross-sheet reference",
        "",
        "METHODOLOGY",
        "  1. Build USD SOFR discount curve from: 1D rate, FOMC step quotes, SOFR futures",
        "     (with convexity adjustment), and par swap rates (bootstrapped annually, ACT/360).",
        "  2. For each calendar day d in the Leg 3 schedule (24-Oct-26 to 24-Apr-31):",
        "     a. Compute the forward 10Y CMS rate: fwd = (DF(d) - DF(d+10Y)) / annuity",
        "     b. Interpolate ATM swaption vol and SABR rho, nu at expiry T = d - val_date.",
        "     c. Calibrate SABR alpha so that SABR-ATM-vol(fwd, T) = ATM vol.",
        "     d. Compute SABR implied vol at K = 4.2%.",
        "     e. Compute digital probability P(S10Y ≤ K) using Black model.",
        "  3. For each of 18 quarterly periods, sum daily probabilities / days = accrual %.",
        "  4. Period coupon PV = Notional × 5.32% × τ(30/360 = 0.25) × accrual% × DF(pay date).",
        "  5. Leg 3 PV = sum of period PVs.",
    ]

    # `enumerate(rows, start=2)` gives us pairs (2, first_string), (3, second_string), ...
    # so we can use `i` as the row number to write to.
    for i, t in enumerate(rows, start=2):
        setcell(ws, i, 1, t, FORMULA_FONT, align=LEFT_ALIGN)


# =============================================================================
# SECTION 6 — BUILD THE MARKETDATA SHEET
# =============================================================================
# This sheet holds ALL the hardcoded inputs. Every other sheet reads from
# here via formulas like =MarketData!$B$5. If you change a number here, the
# entire workbook re-prices.
#
# KEY DESIGN DECISION:
#   As we write each value, we REMEMBER its row number (e.g. `sofr_1d_row = r`)
#   and stash it back into the `data` dictionary (e.g. `data['_md_sofr_1d_cell']`).
#   This way, other sheets that need to reference this value can look up
#   the cell address from the dictionary instead of hardcoding row numbers
#   (which would break if we changed the layout).
#
#   The underscore prefix (_md_...) is a convention: it means "internal —
#   computed during build, not part of the original market data".

def build_market_data(wb, data):
    """Build the MarketData sheet. The only sheet with hardcoded inputs."""
    ws = wb.create_sheet('MarketData')

    # Set column widths. The `for ... in` loop iterates over a list of (col, width) pairs.
    for col, w in [('A', 26), ('B', 16), ('C', 16), ('D', 16)]:
        ws.column_dimensions[col].width = w

    setcell(ws, 1, 1, "Market Data (HARDCODED — change inputs here)", TITLE_FONT)
    setcell(ws, 2, 1, "Only this sheet contains hardcodes. All other sheets are formula-driven.",
            FORMULA_FONT)

    # `r` will be our "current row" pointer. We increment it as we write rows.
    # This is much cleaner than hardcoding row numbers everywhere.
    r = 4

    # ---- Valuation & trade terms ----
    setcell(ws, r, 1, "VALUATION & TRADE TERMS", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    setcell(ws, r, 1, "Valuation date", HEADER_FONT)
    setcell(ws, r, 2, data['val_date'], INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Notional (USD)", HEADER_FONT)
    setcell(ws, r, 2, data['notional'], INPUT_FONT, fmt='#,##0', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Coupon rate", HEADER_FONT)
    setcell(ws, r, 2, data['coupon'], INPUT_FONT, fmt='0.000%', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Barrier K (upper)", HEADER_FONT)
    setcell(ws, r, 2, data['K'], INPUT_FONT, fmt='0.000%', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Leg 3 start (1st accrual period start)", HEADER_FONT)
    setcell(ws, r, 2, data['leg3_start'], INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Leg 3 end (maturity)", HEADER_FONT)
    setcell(ws, r, 2, data['leg3_end'], INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Reference index", HEADER_FONT)
    setcell(ws, r, 2, "10Y SOFR CMS", INPUT_FONT, fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Daycount basis (coupon leg)", HEADER_FONT)
    setcell(ws, r, 2, "30/360", INPUT_FONT, fill=INPUT_FILL); r += 1
    setcell(ws, r, 1, "Daycount basis (curve/annuity)", HEADER_FONT)
    setcell(ws, r, 2, "ACT/360", INPUT_FONT, fill=INPUT_FILL); r += 2

    # ---- 1D SOFR ----
    setcell(ws, r, 1, "1D MONEY MARKET", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    setcell(ws, r, 1, "1D SOFR mid-rate", HEADER_FONT)
    # IMPORTANT: capture the row BEFORE incrementing so we know where sofr_1d lives.
    sofr_1d_row = r
    setcell(ws, r, 2, data['sofr_1d'], INPUT_FONT, fmt='0.0000%', fill=INPUT_FILL); r += 2
    # Stash the cell address in `data` so other functions can refer to it via
    # data['_md_sofr_1d_cell']. The result is a string like 'MarketData!$B$16'.
    data['_md_sofr_1d_cell'] = f'MarketData!$B${sofr_1d_row}'

    # ---- FOMC step quotes (table) ----
    setcell(ws, r, 1, "FOMC STEP QUOTES", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    for h, hv in enumerate(["Start", "End", "Rate"], start=1):
        setcell(ws, r, h, hv, HEADER_FONT, align=CENTER_ALIGN)
    r += 1
    # Remember the first row of the table so we can build a range reference later.
    step_start_row = r
    # Loop over each (start, end, rate) tuple in the list.
    for s, e, rt in data['step_quotes']:
        setcell(ws, r, 1, s, INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL)
        setcell(ws, r, 2, e, INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL)
        setcell(ws, r, 3, rt, INPUT_FONT, fmt='0.0000%', fill=INPUT_FILL)
        r += 1
    step_end_row = r - 1  # we went one past, so subtract 1
    # Stash the (first_row, last_row) range as a tuple
    data['_md_steps_range'] = (step_start_row, step_end_row)
    r += 1  # blank row before next section

    # ---- Futures (table) ----
    setcell(ws, r, 1, "SOFR 3M FUTURES", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    for h, hv in enumerate(["Period start (3rd Wed)", "Period end (3rd Wed)",
                            "Price", "Convexity adj (%)"], start=1):
        setcell(ws, r, h, hv, HEADER_FONT, align=CENTER_ALIGN)
    r += 1
    fut_start_row = r
    for y, m, price, conv in data['futures_raw']:
        # Convert the (year, month) into the actual contract start/end dates.
        ps = third_wed(y, m)
        # The contract ends on the 3rd Wed of the month 3 months later.
        # If the month is December (12), next quarter is March of NEXT year.
        ny, nm = (y + 1, 3) if m == 12 else (y, m + 3)
        pe = third_wed(ny, nm)
        setcell(ws, r, 1, ps, INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL)
        setcell(ws, r, 2, pe, INPUT_FONT, fmt='dd-mmm-yyyy', fill=INPUT_FILL)
        setcell(ws, r, 3, price, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        setcell(ws, r, 4, conv, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        r += 1
    fut_end_row = r - 1
    data['_md_futs_range'] = (fut_start_row, fut_end_row)
    r += 1

    # ---- Swap rates (table) ----
    setcell(ws, r, 1, "PAR SWAP RATES (USD SOFR)", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    for h, hv in enumerate(["Tenor (Y)", "Par rate"], start=1):
        setcell(ws, r, h, hv, HEADER_FONT, align=CENTER_ALIGN)
    r += 1
    swap_start_row = r
    for tenor, rt in data['swap_rates']:
        setcell(ws, r, 1, tenor, INPUT_FONT, fmt='0', fill=INPUT_FILL)
        setcell(ws, r, 2, rt, INPUT_FONT, fmt='0.0000%', fill=INPUT_FILL)
        r += 1
    swap_end_row = r - 1
    data['_md_swaps_range'] = (swap_start_row, swap_end_row)
    r += 1

    # ---- ATM swaption vol (10Y tail) (table) ----
    setcell(ws, r, 1, "ATM SWAPTION VOL (10Y TAIL)", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    for h, hv in enumerate(["Expiry (Y)", "ATM vol"], start=1):
        setcell(ws, r, h, hv, HEADER_FONT, align=CENTER_ALIGN)
    r += 1
    atm_start_row = r
    for T, v in data['atm_10Y']:
        setcell(ws, r, 1, T, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        setcell(ws, r, 2, v, INPUT_FONT, fmt='0.0000%', fill=INPUT_FILL)
        r += 1
    atm_end_row = r - 1
    data['_md_atm_range'] = (atm_start_row, atm_end_row)
    r += 1

    # ---- SABR rho/nu + beta (table) ----
    setcell(ws, r, 1, "SABR PARAMETERS (10Y TAIL)", SUBTITLE_FONT, fill=SECTION_FILL); r += 1
    setcell(ws, r, 1, "Beta", HEADER_FONT)
    setcell(ws, r, 2, data['beta'], INPUT_FONT, fmt='0.00', fill=INPUT_FILL); r += 2
    for h, hv in enumerate(["Expiry (Y)", "Rho", "Nu"], start=1):
        setcell(ws, r, h, hv, HEADER_FONT, align=CENTER_ALIGN)
    r += 1
    sabr_start_row = r
    for T, rho, nu in data['sabr_10Y']:
        setcell(ws, r, 1, T, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        setcell(ws, r, 2, rho, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        setcell(ws, r, 3, nu, INPUT_FONT, fmt='0.0000', fill=INPUT_FILL)
        r += 1
    sabr_end_row = r - 1
    data['_md_sabr_range'] = (sabr_start_row, sabr_end_row)

    # ---- Stash cell references for OTHER sheets to find their inputs ----
    # `str(...)` converts a number to a string so we can concatenate it with text.
    data['_md_beta_cell'] = 'MarketData!$B$' + str(sabr_start_row - 2)
    data['_md_valdate_cell'] = 'MarketData!$B$5'
    data['_md_notional_cell'] = 'MarketData!$B$6'
    data['_md_coupon_cell'] = 'MarketData!$B$7'
    data['_md_K_cell'] = 'MarketData!$B$8'
    data['_md_leg3_start_cell'] = 'MarketData!$B$9'
    data['_md_leg3_end_cell'] = 'MarketData!$B$10'
    # (_md_sofr_1d_cell was set earlier where we wrote the value)


# =============================================================================
# SECTION 7 — BUILD THE YIELDCURVE SHEET
# =============================================================================
# This sheet builds the discount factor curve sequentially:
#
#   1. Start with DF(val_date) = 1.0 by definition.
#   2. From val_date to first FOMC meeting: use the 1D SOFR rate.
#   3. Between FOMC meetings: use the step quotes (each step is a constant rate
#      assumed between two Fed meetings).
#   4. After the last FOMC step: use SOFR 3M futures.
#   5. After the last future: bootstrap from par swap rates (we solve for the
#      DF at each swap maturity that makes the swap fair).
#
# Each step uses:
#     new_DF = previous_DF / (1 + rate × days / 360)
# This is the "money market" / "simple compounding" convention.
#
# The bootstrap formula for swaps is more involved:
#     DF_n = (DF_spot - R_n × Σ_{i=1..n-1} τ_i × DF_i) / (1 + R_n × τ_n)
# where R_n is the n-year par swap rate and τ_i is the year fraction between
# years i-1 and i.

def build_yield_curve(wb, data):
    """Construct the SOFR discount factor curve."""
    ws = wb.create_sheet('YieldCurve')

    # Set column widths individually
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 50

    setcell(ws, 1, 1, "Yield Curve Construction", TITLE_FONT)
    setcell(ws, 2, 1, "Sequential bootstrap: MM → FOMC steps → futures → swaps. "
                       "Each DF computed from previous DF and the relevant rate over a period.",
            FORMULA_FONT, align=LEFT_ALIGN)
    ws.row_dimensions[2].height = 30  # taller row for the wrap

    # Column headers
    headers = ["Pillar #", "Source", "Period start", "Period end", "Rate / Fwd",
               "Discount factor", "Formula explanation"]
    for i, h in enumerate(headers, start=1):
        setcell(ws, 4, i, h, HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)

    r = 5  # data starts at row 5

    # ----- Pillar 1: val_date with DF = 1 -----
    setcell(ws, r, 1, 1, FORMULA_FONT)
    setcell(ws, r, 2, "Valuation date", FORMULA_FONT)
    # Formulas are STRINGS that start with '='. The f"..." string interpolates
    # data['_md_valdate_cell'] = 'MarketData!$B$5' so we get '=MarketData!$B$5'.
    setcell(ws, r, 3, f"={data['_md_valdate_cell']}", XREF_FONT, fmt='dd-mmm-yyyy')
    setcell(ws, r, 4, f"={data['_md_valdate_cell']}", XREF_FONT, fmt='dd-mmm-yyyy')
    setcell(ws, r, 5, "", FORMULA_FONT)
    setcell(ws, r, 6, 1.0, FORMULA_FONT, fmt='0.000000000000')
    setcell(ws, r, 7, "DF(val_date) = 1 by definition", FORMULA_FONT, align=LEFT_ALIGN)
    r += 1

    # ----- Pillar 2: end of 1D bridge (= first step start) -----
    # Get the (first_row, last_row) range for steps from data
    step_s_row, step_e_row = data['_md_steps_range']
    # Build a cell reference to the FIRST step start (column A of first step row)
    first_step_start = f"MarketData!$A${step_s_row}"
    setcell(ws, r, 1, 2, FORMULA_FONT)
    setcell(ws, r, 2, "1D SOFR bridge", FORMULA_FONT)
    setcell(ws, r, 3, f"=C{r-1}", FORMULA_FONT, fmt='dd-mmm-yyyy')   # start = previous end
    setcell(ws, r, 4, f"={first_step_start}", XREF_FONT, fmt='dd-mmm-yyyy')
    setcell(ws, r, 5, f"={data['_md_sofr_1d_cell']}", XREF_FONT, fmt='0.0000%')
    # The compounding formula: new_DF = prev_DF / (1 + rate × days / 360)
    setcell(ws, r, 6, f"=F{r-1}/(1+E{r}*(D{r}-C{r})/360)", FORMULA_FONT,
            fmt='0.000000000000')
    setcell(ws, r, 7, "DF = prev_DF / (1 + rate × days/360)", FORMULA_FONT, align=LEFT_ALIGN)
    r += 1

    # ----- FOMC step pillars: one row per step quote -----
    n_steps = step_e_row - step_s_row + 1
    for i in range(n_steps):
        md_row = step_s_row + i
        setcell(ws, r, 1, i + 3, FORMULA_FONT)
        setcell(ws, r, 2, f"FOMC step {i+1}", FORMULA_FONT)
        # Each step's start, end, rate come from the corresponding MarketData row
        setcell(ws, r, 3, f"=MarketData!$A${md_row}", XREF_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 4, f"=MarketData!$B${md_row}", XREF_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 5, f"=MarketData!$C${md_row}", XREF_FONT, fmt='0.0000%')
        setcell(ws, r, 6, f"=F{r-1}/(1+E{r}*(D{r}-C{r})/360)", FORMULA_FONT,
                fmt='0.000000000000')
        if i == 0:
            setcell(ws, r, 7, "Each FOMC step extends the DF curve to next FOMC date",
                    FORMULA_FONT, align=LEFT_ALIGN)
        r += 1

    # ----- Futures pillars -----
    fut_s_row, fut_e_row = data['_md_futs_range']
    n_futs = fut_e_row - fut_s_row + 1
    for i in range(n_futs):
        md_row = fut_s_row + i
        setcell(ws, r, 1, r - 4, FORMULA_FONT)
        setcell(ws, r, 2, f"Future #{i+1}", FORMULA_FONT)
        # The effective start of a future is MAX(previous curve end, future's standard start).
        # This handles the case where the last FOMC step overlaps with the start of
        # the first future.
        setcell(ws, r, 3, f"=MAX(D{r-1}, MarketData!$A${md_row})", FORMULA_FONT,
                fmt='dd-mmm-yyyy')
        setcell(ws, r, 4, f"=MarketData!$B${md_row}", XREF_FONT, fmt='dd-mmm-yyyy')
        # Forward rate from future price: (100 - price)/100 - convexity_adj/100
        setcell(ws, r, 5, f"=(100-MarketData!$C${md_row})/100-MarketData!$D${md_row}/100",
                FORMULA_FONT, fmt='0.0000%')
        setcell(ws, r, 6, f"=F{r-1}/(1+E{r}*(D{r}-C{r})/360)", FORMULA_FONT,
                fmt='0.000000000000')
        if i == 0:
            setcell(ws, r, 7, "Future fwd = (100 - price)/100 - convexity adj. "
                              "Overlap with last step handled by MAX(prev end, future start).",
                    FORMULA_FONT, align=LEFT_ALIGN)
            ws.row_dimensions[r].height = 28
        r += 1

    last_fut_row = r - 1  # remember where futures end

    # ----- Spot date helper (placed off to the side in columns N-O) -----
    # IMPORTANT: we put the spot helper in DIFFERENT COLUMNS than the main pillar
    # list, because spot_date < first_step_start (it's only val_date + 2 days),
    # and if we put it in the main column D list it would break the ascending-
    # date sort that MATCH(...,1) requires.
    setcell(ws, 4, 14, "Spot helper", HEADER_FONT, fill=SECTION_FILL)
    setcell(ws, 5, 14, "Spot date", HEADER_FONT)
    setcell(ws, 5, 15, f"={data['_md_valdate_cell']}+2", FORMULA_FONT, fmt='dd-mmm-yyyy')
    setcell(ws, 6, 14, "DF(spot)", HEADER_FONT)
    # Spot is 2 days after val_date, in the 1D bridge zone, so:
    # DF(spot) = 1 / (1 + sofr_1d × 2 / 360)
    setcell(ws, 6, 15, f"=1/(1+{data['_md_sofr_1d_cell']}*2/360)",
            FORMULA_FONT, fmt='0.000000000000')
    setcell(ws, 7, 14, "Note:", FORMULA_FONT)
    setcell(ws, 7, 15, "Spot = val_date + 2 days; DF = 1/(1+sofr_1d·2/360)",
            FORMULA_FONT, align=LEFT_ALIGN)
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 18

    # Stash spot cell addresses for the bootstrap formulas
    spot_date_cell = "$O$5"
    spot_df_cell = "$O$6"
    data['_yc_spot_date_cell'] = spot_date_cell
    data['_yc_spot_df_cell'] = spot_df_cell

    # ===== Annual anniversaries helper table (cols I-M) =====
    # For each year n from 1 to max_tenor we set up a row showing:
    #   - the date (spot + n years)
    #   - the par rate at year n (interpolated linearly when needed)
    #   - the year fraction τ_n (ACT/360 from year n-1 to year n)
    #   - the discount factor DF_n
    #
    # Years 1, 2, 3: DF is log-linear interpolated from the futures range
    #                (since the futures cover that horizon).
    # Years 4+ : DF is bootstrapped from the par swap rate.

    swap_s_row, swap_e_row = data['_md_swaps_range']
    swap_tenors_list = [t for t, _ in data['swap_rates']]  # list comprehension
    max_tenor = max(swap_tenors_list)        # built-in max
    min_swap_tenor = min(swap_tenors_list)   # built-in min (will be 4)

    setcell(ws, 4, 9, "Year #", HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)
    setcell(ws, 4, 10, "Pay date", HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)
    setcell(ws, 4, 11, "Par rate (interp)", HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)
    setcell(ws, 4, 12, "τ_n (ACT/360)", HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)
    setcell(ws, 4, 13, "DF_n", HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)

    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 11
    ws.column_dimensions['M'].width = 16

    annual_first_row = 6
    spot_cell = spot_date_cell

    # Build range references to the swap tenor and rate columns in MarketData
    md_tenor_range = f"MarketData!$A${swap_s_row}:$A${swap_e_row}"
    md_rate_range = f"MarketData!$B${swap_s_row}:$B${swap_e_row}"

    last_fut_pillar_row = last_fut_row

    # Step 1: write the date and par rate (interp) for every year
    for n in range(1, max_tenor + 1):
        rr = annual_first_row + (n - 1)  # row in the helper table
        setcell(ws, rr, 9, n, FORMULA_FONT)
        # Pay date = EDATE(spot, 12*n) → n years after spot
        setcell(ws, rr, 10, f"=EDATE({spot_cell},{12*n})", FORMULA_FONT, fmt='dd-mmm-yyyy')

        # Par rate: linear interp from the swap rate table
        if n < min_swap_tenor:
            # Years 1, 2, 3 don't need a par rate (not bootstrapped via swap)
            setcell(ws, rr, 11, "", FORMULA_FONT)
        else:
            # The IF handles the edge case where n is exactly the last tenor.
            # The big chained INDEX/MATCH does standard linear interpolation:
            #   slope = (rate[next] - rate[curr]) / (tenor[next] - tenor[curr])
            #   value = rate[curr] + (n - tenor[curr]) × slope
            par_formula = (
                f"=IF(I{rr}>=INDEX({md_tenor_range},ROWS({md_tenor_range})),"
                f"INDEX({md_rate_range},ROWS({md_rate_range})),"
                f"INDEX({md_rate_range},MATCH(I{rr},{md_tenor_range},1))"
                f"+(I{rr}-INDEX({md_tenor_range},MATCH(I{rr},{md_tenor_range},1)))/"
                f"(INDEX({md_tenor_range},MATCH(I{rr},{md_tenor_range},1)+1)"
                f"-INDEX({md_tenor_range},MATCH(I{rr},{md_tenor_range},1)))*"
                f"(INDEX({md_rate_range},MATCH(I{rr},{md_tenor_range},1)+1)"
                f"-INDEX({md_rate_range},MATCH(I{rr},{md_tenor_range},1))))"
            )
            setcell(ws, rr, 11, par_formula, FORMULA_FONT, fmt='0.0000%')

        # τ_n: ACT/360 year fraction
        if n == 1:
            setcell(ws, rr, 12, f"=(J{rr}-{spot_cell})/360", FORMULA_FONT, fmt='0.0000')
        else:
            setcell(ws, rr, 12, f"=(J{rr}-J{rr-1})/360", FORMULA_FONT, fmt='0.0000')

    # Step 2: fill in DFs.
    # Years 1, 2, 3: log-linear interpolation from the futures-range pillars
    for n in range(1, min_swap_tenor):
        rr = annual_first_row + (n - 1)
        # The log-linear interpolation formula. Let me walk through what's
        # happening:
        #
        # Goal: find DF at date J_rr by interpolating between two known pillars
        #       in the curve we've built so far (rows 5 to last_fut_pillar_row).
        #
        # MATCH(J_rr, dates, 1)
        #   returns the position of the LAST date <= J_rr.
        #   (works because the date list is sorted ascending)
        # INDEX(F-list, pos) returns the DF at that position.
        # INDEX(F-list, pos+1) returns the next DF.
        # INDEX(D-list, pos) and (pos+1) return the bracketing dates.
        # We then compute:
        #     fraction = (target - date_lo) / (date_hi - date_lo)
        #     log_DF   = log(DF_lo) + fraction × (log(DF_hi) - log(DF_lo))
        #     DF       = exp(log_DF)
        interp = (
            f"=EXP(LN(INDEX($F$5:$F${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)))"
            f"+(J{rr}-INDEX($D$5:$D${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)))/"
            f"(INDEX($D$5:$D${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)+1)"
            f"-INDEX($D$5:$D${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)))*"
            f"(LN(INDEX($F$5:$F${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)+1))"
            f"-LN(INDEX($F$5:$F${last_fut_pillar_row},"
            f"MATCH(J{rr},$D$5:$D${last_fut_pillar_row},1)))))"
        )
        setcell(ws, rr, 13, interp, FORMULA_FONT, fmt='0.000000000000')

    # Years 4..max_tenor: bootstrap.
    # DF_n = (DF_spot - R_n × Σ_{i=1..n-1} τ_i × DF_i) / (1 + R_n × τ_n)
    for n in range(min_swap_tenor, max_tenor + 1):
        rr = annual_first_row + (n - 1)
        # Build the annuity sum dynamically: τ_1·DF_1 + τ_2·DF_2 + ... + τ_{n-1}·DF_{n-1}
        sum_terms = []
        for i in range(1, n):
            ii_row = annual_first_row + (i - 1)
            sum_terms.append(f"L{ii_row}*M{ii_row}")
        # `"+".join([...])` glues list items together with "+" between them
        sum_expr = "+".join(sum_terms)
        bootstrap = (
            f"=({data['_yc_spot_df_cell']}-K{rr}*({sum_expr}))"
            f"/(1+K{rr}*L{rr})"
        )
        setcell(ws, rr, 13, bootstrap, FORMULA_FONT, fmt='0.000000000000')

    # Step 3: copy bootstrapped pillars into the MAIN pillar list (cols D, F).
    # This is what DailyDFs will use to interpolate.
    for n in range(min_swap_tenor, max_tenor + 1):
        rr_helper = annual_first_row + (n - 1)
        setcell(ws, r, 1, r - 4, FORMULA_FONT)
        setcell(ws, r, 2, f"Bootstrap {n}Y", FORMULA_FONT)
        setcell(ws, r, 3, f"={spot_cell}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 4, f"=J{rr_helper}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 5, f"=K{rr_helper}", XREF_FONT, fmt='0.0000%')
        setcell(ws, r, 6, f"=M{rr_helper}", FORMULA_FONT, fmt='0.000000000000')
        if n == min_swap_tenor:
            setcell(ws, r, 7, "DF_n = (DF_spot − R_n × Σ τᵢDFᵢ) / (1 + R_n × τ_n). "
                              "Years not in MarketData (e.g. 13Y, 14Y) get par rate via linear interp.",
                    FORMULA_FONT, align=LEFT_ALIGN)
            ws.row_dimensions[r].height = 28
        r += 1

    # Save the last row of the pillar list for other sheets
    last_pillar_row = r - 1
    data['_yc_last_pillar_row'] = last_pillar_row

    # Print a summary row
    r += 1
    setcell(ws, r, 1, "Curve construction complete. Final DF at end of last swap pillar:",
            HEADER_FONT); r += 1
    setcell(ws, r, 1, f"DF at last pillar ({max_tenor}Y swap maturity):", FORMULA_FONT)
    setcell(ws, r, 2, f"=F{last_pillar_row}", FORMULA_FONT, fmt='0.000000000000')


# =============================================================================
# SECTION 8 — BUILD THE DAILYDFs SHEET
# =============================================================================
# This sheet has ONE ROW per calendar day from val_date to (leg3_end + 10 years).
# Each row holds the discount factor at that day, computed by log-linear
# interpolation on the YieldCurve pillars.
#
# Why pre-compute this? Because DailyFixings needs DF lookups at ~18,000 dates
# (1640 days × 11 DFs per day). If we did log-linear interp inline there, each
# lookup would be 6 nested INDEX/MATCH calls and the workbook would become slow.
# By pre-computing daily DFs once here, the DailyFixings lookups become simple
# INDEX-by-row-offset operations — O(1) lookups, much faster.

def build_daily_dfs(wb, data):
    """Pre-compute the DF at every calendar day."""
    ws = wb.create_sheet('DailyDFs')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18

    setcell(ws, 1, 1, "Daily DF Table (log-linear interp from YieldCurve pillars)", TITLE_FONT)
    setcell(ws, 2, 1, "DF at every day from val_date to maturity + 10 years. "
                       "Used as fast lookup table for DailyFixings.",
            FORMULA_FONT, align=LEFT_ALIGN)
    ws.row_dimensions[2].height = 26

    headers = ["Day idx", "Date", "DF (log-linear interp)"]
    for i, h in enumerate(headers, start=1):
        setcell(ws, 4, i, h, HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)

    val_d = data['val_date']
    # We need DFs from val_date all the way to (leg3_end + 10 years + a buffer)
    end_d = add_months(data['leg3_end'], 12 * 10 + 1)  # 121 months past leg3_end
    n_days = (end_d - val_d).days + 1

    last_pillar = data['_yc_last_pillar_row']

    # Generate one row per day. This produces a few thousand rows.
    for i in range(n_days):
        r = i + 5  # data starts at row 5
        setcell(ws, r, 1, i, FORMULA_FONT)
        # Date = val_date + i days
        setcell(ws, r, 2, f"={data['_md_valdate_cell']}+{i}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        # Log-linear interpolation formula. The IFERROR wraps it so that if any
        # date falls outside the pillar range, we return 1 (DF=1, no interest).
        # In practice this never triggers for valid trade dates.
        formula = (
            f"=IFERROR("
            f"EXP("
            f"LN(INDEX(YieldCurve!$F$5:$F${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1)))"
            f"+(B{r}-INDEX(YieldCurve!$D$5:$D${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1)))/"
            f"(INDEX(YieldCurve!$D$5:$D${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1)+1)"
            f"-INDEX(YieldCurve!$D$5:$D${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1)))*"
            f"(LN(INDEX(YieldCurve!$F$5:$F${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1)+1))"
            f"-LN(INDEX(YieldCurve!$F$5:$F${last_pillar},"
            f"MATCH(B{r}, YieldCurve!$D$5:$D${last_pillar}, 1))))"
            f"),1)"
        )
        setcell(ws, r, 3, formula, FORMULA_FONT, fmt='0.0000000000')

    data['_dfs_first_row'] = 5
    data['_dfs_last_row'] = 5 + n_days - 1


# =============================================================================
# SECTION 9 — HELPER FORMULA BUILDERS
# =============================================================================
# These three Python functions don't write to Excel — they BUILD formula
# strings that we plug into Excel cells. They make the code cleaner: instead
# of repeating a 200-character formula 1640 times, we call a function that
# substitutes the cell references for us.

def atm_interp_formula(T_range, V_range, T_cell):
    """
    Build a linear interpolation formula.

    Given a table of (T, V) pairs (T in T_range, V in V_range), interpolate V
    at the value of T_cell. Clamps at endpoints if T_cell is out of range.

    Standard linear interp:
        V(T) = V_lo + (T - T_lo) / (T_hi - T_lo) × (V_hi - V_lo)
    """
    return (
        f"=IF({T_cell}<=INDEX({T_range},1),INDEX({V_range},1),"
        f"IF({T_cell}>=INDEX({T_range},ROWS({T_range})),INDEX({V_range},ROWS({V_range})),"
        f"INDEX({V_range},MATCH({T_cell},{T_range},1))+"
        f"({T_cell}-INDEX({T_range},MATCH({T_cell},{T_range},1)))/"
        f"(INDEX({T_range},MATCH({T_cell},{T_range},1)+1)-INDEX({T_range},MATCH({T_cell},{T_range},1)))*"
        f"(INDEX({V_range},MATCH({T_cell},{T_range},1)+1)-INDEX({V_range},MATCH({T_cell},{T_range},1)))))"
    )


def sabr_alpha_update_formula(alpha_cell, f_cell, T_cell, beta_cell,
                              rho_cell, nu_cell, sigma_atm_cell):
    """
    Build the SABR α calibration update formula.

    MATH:
    The SABR ATM lognormal vol is approximately:
        σ_ATM = (α/f^(1-β)) × (1 + T × [(1-β)²/24 × α²/f^(2-2β)
                                         + ρβνα/(4f^(1-β))
                                         + (2-3ρ²)/24 × ν²])

    We need to find α such that this equals the market ATM vol. We use
    a simple iteration: α_new = α_old × σ_market / σ_computed(α_old).
    Starting from α_0 = σ_market × f^(1-β) and iterating 2-3 times gives
    convergence to ~6 decimal places.
    """
    a, f, T, b, rho, nu, sa = alpha_cell, f_cell, T_cell, beta_cell, rho_cell, nu_cell, sigma_atm_cell
    sabr_atm = (
        f"({a}/{f}^(1-{b}))*"
        f"(1+{T}*"
        f"((1-{b})^2/24*{a}^2/{f}^(2-2*{b})"
        f"+{rho}*{b}*{nu}*{a}/(4*{f}^(1-{b}))"
        f"+(2-3*{rho}^2)/24*{nu}^2))"
    )
    return f"={a}*{sa}/({sabr_atm})"


def sabr_vol_at_K_formula(f_cell, K_cell, T_cell, alpha_cell, beta_cell, rho_cell, nu_cell):
    """
    Build the SABR implied lognormal vol formula at strike K.

    This is the Hagan 2002 formula. It's long but mechanical.
    The IF at the front handles the case f ≈ K where the standard formula has
    a 0/0; in that case we use the ATM formula directly.
    """
    a, K, T, b, rho, nu, f = alpha_cell, K_cell, T_cell, beta_cell, rho_cell, nu_cell, f_cell
    log_fK = f"LN({f}/{K})"
    fKb = f"({f}*{K})^((1-{b})/2)"
    z = f"({nu}/{a})*{fKb}*{log_fK}"
    xz = f"LN((SQRT(1-2*{rho}*({z})+({z})^2)+({z})-{rho})/(1-{rho}))"
    pre = (
        f"({a}/("
        f"{fKb}*(1+(1-{b})^2/24*({log_fK})^2+(1-{b})^4/1920*({log_fK})^4)))"
    )
    cor = (
        f"(1+{T}*"
        f"((1-{b})^2/24*{a}^2/({f}*{K})^(1-{b})"
        f"+{rho}*{b}*{nu}*{a}/(4*{fKb})"
        f"+(2-3*{rho}^2)/24*{nu}^2))"
    )
    atm_branch = (
        f"({a}/{f}^(1-{b}))*"
        f"(1+{T}*"
        f"((1-{b})^2/24*{a}^2/{f}^(2-2*{b})"
        f"+{rho}*{b}*{nu}*{a}/(4*{f}^(1-{b}))"
        f"+(2-3*{rho}^2)/24*{nu}^2))"
    )
    return f"=IF(ABS({f}-{K})<1E-10,{atm_branch},{pre}*({z}/({xz}))*{cor})"


# =============================================================================
# SECTION 10 — BUILD THE DAILYFIXINGS SHEET
# =============================================================================
# This is the heart of the workbook. ONE ROW per calendar day in Leg 3
# (~1640 rows). Each row computes everything needed for that day's accrual:
#
#   - DF at the fixing day
#   - DF at fix_day + 1Y, 2Y, ..., 10Y  (used to compute the forward 10Y CMS)
#   - τ_i year fractions
#   - Annuity = Σ τ × DF
#   - Forward 10Y CMS = (DF_0 - DF_10Y) / Annuity
#   - Time to expiry T (in years)
#   - ATM vol, SABR ρ, SABR ν (all interpolated from MarketData tables)
#   - SABR α (calibrated to match the ATM vol)
#   - SABR vol at K = 4.2%
#   - d2 of the Black model
#   - Digital probability P(S ≤ K) = N(-d2)

def build_daily_fixings(wb, data):
    """Compute daily forwards, SABR vols, and digital probabilities."""
    ws = wb.create_sheet('DailyFixings')

    # ---- column definitions ----
    # `cols_def` is a list of (letter, header, width) triples that describes
    # the layout. We'll loop over it to set widths and headers.
    cols_def = [
        ('A', 'Idx', 6),
        ('B', 'Date', 11),
        ('C', 'T (years)', 9),
        ('D', 'DF(d)', 11),
    ]
    # DF at d + 1Y .. d + 10Y go in columns E..N (10 columns)
    for i in range(1, 11):
        col = get_column_letter(4 + i)  # 4+1=5 → 'E', ..., 4+10=14 → 'N'
        cols_def.append((col, f'DF(d+{i}Y)', 11))
    # τ_1 .. τ_10 go in columns O..X (10 columns)
    for i in range(1, 11):
        col = get_column_letter(14 + i)
        cols_def.append((col, f'τ_{i}', 8))
    # Remaining columns
    extra = [
        ('Y',  'Annuity', 11),
        ('Z',  'fwd 10Y', 11),
        ('AA', 'ATM σ', 9),
        ('AB', 'rho', 9),
        ('AC', 'nu', 9),
        ('AD', 'α (init)', 9),
        ('AE', 'α iter 1', 9),
        ('AF', 'α iter 2', 9),
        ('AG', 'σ at K', 9),
        ('AH', 'd2', 9),
        ('AI', 'P(S≤K)', 10),
    ]
    cols_def += extra

    # Set widths and headers
    for col, hdr, w in cols_def:
        ws.column_dimensions[col].width = w

    setcell(ws, 1, 1, "Daily Fixings (one row per calendar day in Leg 3)", TITLE_FONT)
    setcell(ws, 2, 1, "For each day d: compute forward 10Y CMS from DFs, then SABR vol at K, "
                       "then digital probability P(S₁₀ᵧ ≤ 4.2%) using Black model.",
            FORMULA_FONT, align=LEFT_ALIGN)
    ws.row_dimensions[2].height = 26

    for i, (col, hdr, _) in enumerate(cols_def, start=1):
        setcell(ws, 4, i, hdr, HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)

    # ---- generate one row per calendar day in Leg 3 ----
    val_d = data['val_date']
    s = data['leg3_start']
    e = data['leg3_end']
    n_days = (e - s).days
    val_offset = (s - val_d).days  # offset in days from val_date to leg 3 start

    # Stash cell references we use repeatedly
    K_cell = data['_md_K_cell']
    beta_cell = data['_md_beta_cell']
    val_cell = data['_md_valdate_cell']
    atm_s, atm_e = data['_md_atm_range']
    sabr_s, sabr_e = data['_md_sabr_range']

    data['_df_first_row'] = 5
    data['_df_last_row'] = 5 + n_days - 1

    for i in range(n_days):
        r = i + 5  # row index (data starts at row 5)

        # Index
        setcell(ws, r, 1, i + 1, FORMULA_FONT)
        # Date = val_date + (val_offset + i) — this is calendar day i of Leg 3
        setcell(ws, r, 2, f"={val_cell}+{val_offset + i}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        # Time to expiry in years. Cap at 1/365.25 to avoid division by 0 for very near dates.
        setcell(ws, r, 3, f"=MAX((B{r}-{val_cell})/365.25, 1/365.25)",
                FORMULA_FONT, fmt='0.0000')
        # DF(d) — look up in DailyDFs by computing row offset
        # DailyDFs row for date X is (X - val_date) + 5, so:
        setcell(ws, r, 4, f"=INDEX(DailyDFs!$C:$C, B{r}-{val_cell}+5)",
                FORMULA_FONT, fmt='0.0000000000')

        # DF at d + k years, for k = 1..10
        for k in range(1, 11):
            col = 4 + k  # E=5 for k=1, ..., N=14 for k=10
            setcell(ws, r, col,
                    f"=INDEX(DailyDFs!$C:$C, EDATE(B{r},{12*k})-{val_cell}+5)",
                    FORMULA_FONT, fmt='0.0000000000')

        # τ_k = (date(d+kY) - date(d+(k-1)Y)) / 360, ACT/360 daycount
        for k in range(1, 11):
            col = 14 + k  # O=15 for k=1
            if k == 1:
                setcell(ws, r, col, f"=(EDATE(B{r},12)-B{r})/360",
                        FORMULA_FONT, fmt='0.0000')
            else:
                setcell(ws, r, col, f"=(EDATE(B{r},{12*k})-EDATE(B{r},{12*(k-1)}))/360",
                        FORMULA_FONT, fmt='0.0000')

        # Annuity = Σ τ_k × DF(d+kY).
        # SUMPRODUCT multiplies two arrays element-wise and sums.
        setcell(ws, r, 25, f"=SUMPRODUCT(E{r}:N{r}, O{r}:X{r})",
                FORMULA_FONT, fmt='0.0000000000')

        # Forward 10Y = (DF(d) - DF(d+10Y)) / annuity
        # DF(d) is in col D, DF(d+10Y) is in col N.
        setcell(ws, r, 26, f"=(D{r}-N{r})/Y{r}", FORMULA_FONT, fmt='0.000000%')

        # ATM vol interpolation
        atm_range_T = f"MarketData!$A${atm_s}:$A${atm_e}"
        atm_range_V = f"MarketData!$B${atm_s}:$B${atm_e}"
        setcell(ws, r, 27, atm_interp_formula(atm_range_T, atm_range_V, f"C{r}"),
                FORMULA_FONT, fmt='0.0000%')

        # rho, nu interpolation
        sabr_range_T = f"MarketData!$A${sabr_s}:$A${sabr_e}"
        rho_range = f"MarketData!$B${sabr_s}:$B${sabr_e}"
        nu_range  = f"MarketData!$C${sabr_s}:$C${sabr_e}"
        setcell(ws, r, 28, atm_interp_formula(sabr_range_T, rho_range, f"C{r}"),
                FORMULA_FONT, fmt='0.0000')
        setcell(ws, r, 29, atm_interp_formula(sabr_range_T, nu_range, f"C{r}"),
                FORMULA_FONT, fmt='0.0000')

        # SABR alpha calibration: 2 iterations
        # α_0 = σ_ATM × f^(1-β)  (rough first guess)
        setcell(ws, r, 30, f"=AA{r}*Z{r}^(1-{beta_cell})", FORMULA_FONT, fmt='0.0000')
        # Iter 1 (uses α_0 from column AD)
        setcell(ws, r, 31, sabr_alpha_update_formula(f"AD{r}", f"Z{r}", f"C{r}",
                                                     beta_cell, f"AB{r}", f"AC{r}", f"AA{r}"),
                FORMULA_FONT, fmt='0.0000')
        # Iter 2 (uses α_1 from column AE) — this is our final α
        setcell(ws, r, 32, sabr_alpha_update_formula(f"AE{r}", f"Z{r}", f"C{r}",
                                                     beta_cell, f"AB{r}", f"AC{r}", f"AA{r}"),
                FORMULA_FONT, fmt='0.0000')

        # SABR vol at strike K, using the converged α from column AF
        setcell(ws, r, 33, sabr_vol_at_K_formula(f"Z{r}", K_cell, f"C{r}",
                                                  f"AF{r}", beta_cell, f"AB{r}", f"AC{r}"),
                FORMULA_FONT, fmt='0.0000%')

        # Black model d2 = (ln(f/K) - 0.5·σ²·T) / (σ·√T)
        setcell(ws, r, 34, f"=(LN(Z{r}/{K_cell})-0.5*AG{r}^2*C{r})/(AG{r}*SQRT(C{r}))",
                FORMULA_FONT, fmt='0.0000')

        # P(S ≤ K) = 1 - N(d2)  where N is the standard normal CDF
        # NORMSDIST is the older Excel name (works in both Excel and LibreOffice;
        # the newer name NORM.S.DIST can fail in LibreOffice without _xlfn prefix).
        setcell(ws, r, 35, f"=1-NORMSDIST(AH{r})", FORMULA_FONT, fmt='0.0000%')


# =============================================================================
# SECTION 11 — BUILD THE PERIODSUMMARY SHEET
# =============================================================================
# This sheet aggregates the daily probabilities into 18 quarterly coupons.
# Each period has:
#   - Number of calendar days in the period
#   - Accrual % = average daily probability over the period
#   - DF(pay date) = discount factor at the payment date
#   - Coupon PV = Notional × coupon rate × 0.25 × accrual % × DF(pay)
#
# (0.25 is the 30/360 year fraction for one quarter, since 90/360 = 0.25.)

def build_period_summary(wb, data):
    """Aggregate daily fixings into 18 quarterly coupon PVs."""
    ws = wb.create_sheet('PeriodSummary')

    for col, w in [('A', 6), ('B', 12), ('C', 12), ('D', 12), ('E', 8), ('F', 10),
                   ('G', 10), ('H', 12), ('I', 12), ('J', 14)]:
        ws.column_dimensions[col].width = w

    setcell(ws, 1, 1, "Period Summary — aggregation of daily fixings into 18 quarterly coupons",
            TITLE_FONT)
    setcell(ws, 2, 1, "Each row = one quarterly coupon. Accrual % = average daily probability.  "
                       "PV = N × cpn × τ(30/360=0.25) × accrual% × DF(pay).",
            FORMULA_FONT, align=LEFT_ALIGN)
    ws.row_dimensions[2].height = 26

    headers = ["#", "Start", "End", "Pay date", "Days", "Idx start (DailyFixings)",
               "Idx end (DailyFixings)", "Accrual %", "DF(pay)", "Coupon PV (USD)"]
    for i, h in enumerate(headers, start=1):
        setcell(ws, 4, i, h, HEADER_FONT, fill=SECTION_FILL, align=CENTER_ALIGN)

    # Build the schedule of quarterly periods.
    # cur = current period start; we advance by 3 months each loop.
    s = data['leg3_start']
    e = data['leg3_end']
    periods = []
    cur = s
    while cur < e:
        nxt = min(add_months(cur, 3), e)   # never overshoot maturity
        periods.append((cur, nxt))
        cur = nxt

    val_cell = data['_md_valdate_cell']
    leg3_start_cell = data['_md_leg3_start_cell']
    notional_cell = data['_md_notional_cell']
    coupon_cell = data['_md_coupon_cell']

    cum_idx = 0  # running count of days into Leg 3
    for i, (ps, pe) in enumerate(periods):
        r = 5 + i
        days = (pe - ps).days
        idx_start = cum_idx + 1
        idx_end = cum_idx + days

        setcell(ws, r, 1, i + 1, FORMULA_FONT)
        setcell(ws, r, 2, f"={leg3_start_cell}+{cum_idx}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 3, f"={leg3_start_cell}+{cum_idx + days}", FORMULA_FONT, fmt='dd-mmm-yyyy')
        setcell(ws, r, 4, f"=C{r}", FORMULA_FONT, fmt='dd-mmm-yyyy')  # pay = end
        setcell(ws, r, 5, f"=C{r}-B{r}", FORMULA_FONT, fmt='0')
        setcell(ws, r, 6, idx_start, FORMULA_FONT, fmt='0')
        setcell(ws, r, 7, idx_end, FORMULA_FONT, fmt='0')

        # Accrual % = average of daily probabilities in DailyFixings column AI.
        # Row for daily index j is 5 + j - 1 = j + 4.
        first_row = idx_start + 4
        last_row = idx_end + 4
        setcell(ws, r, 8, f"=AVERAGE(DailyFixings!$AI${first_row}:$AI${last_row})",
                FORMULA_FONT, fmt='0.00%')

        # DF at pay date — look up in DailyDFs
        setcell(ws, r, 9, f"=INDEX(DailyDFs!$C:$C, C{r}-{val_cell}+5)",
                FORMULA_FONT, fmt='0.000000')

        # Coupon PV = N × cpn × 0.25 × accrual% × DF(pay)
        setcell(ws, r, 10, f"={notional_cell}*{coupon_cell}*0.25*H{r}*I{r}",
                FORMULA_FONT, fmt='#,##0.00')

        cum_idx += days

    last_period_row = 5 + len(periods) - 1
    data['_ps_last_row'] = last_period_row
    data['_n_periods'] = len(periods)

    # Total
    r = last_period_row + 2
    setcell(ws, r, 1, "TOTAL Leg 3 PV", HEADER_FONT, fill=RESULT_FILL)
    setcell(ws, r, 10, f"=SUM(J5:J{last_period_row})", HEADER_FONT, fmt='#,##0.00',
            fill=RESULT_FILL)
    # Stash the total cell address for the Valuation sheet
    data['_total_pv_cell'] = f"PeriodSummary!$J${r}"

    r += 1
    setcell(ws, r, 1, "% of notional", HEADER_FONT)
    setcell(ws, r, 10, f"=J{r-1}/{notional_cell}", FORMULA_FONT, fmt='0.000%')


# =============================================================================
# SECTION 12 — BUILD THE VALUATION SHEET (step-by-step walkthrough)
# =============================================================================

def build_valuation(wb, data):
    """A beginner-friendly walkthrough of the 8 steps in the calculation."""
    ws = wb.create_sheet('Valuation')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40

    setcell(ws, 1, 2, f"Step-by-Step Valuation — {data['label']}", TITLE_FONT)
    setcell(ws, 2, 2, "Left column = explanation for beginners.  "
                       "Right column = formula reference / result.",
            FORMULA_FONT, align=LEFT_ALIGN)

    r = 4
    setcell(ws, r, 2, "STEP", HEADER_FONT, fill=SECTION_FILL)
    setcell(ws, r, 3, "Cell / value", HEADER_FONT, fill=SECTION_FILL)
    setcell(ws, r, 4, "Notes", HEADER_FONT, fill=SECTION_FILL)
    r += 1

    # Each step is a 4-tuple: (title, explanation, cell_reference, side_note)
    steps = [
        ("Step 1 — Read trade & market inputs",
         "All inputs live on the MarketData sheet. The valuation date, notional, coupon, "
         "barrier K=4.2%, accrual schedule, and the rate, vol, and SABR data are entered "
         "there. Everything else in this workbook is a formula.",
         data['_md_valdate_cell'], "Valuation date"),

        ("Step 2 — Build the discount curve",
         "We need DF(t) at every future date for discounting cash flows and computing "
         "forward rates. Build it sequentially:\n"
         "  • Use the 1D SOFR rate from val_date to the first FOMC meeting.\n"
         "  • Use each FOMC step-quote rate over the period between consecutive Fed meetings.\n"
         "  • Use each SOFR 3M future, converted to a forward rate via "
         "(100 - price)/100 - convexity_adj, over its 3-month period.\n"
         "  • Beyond the last future, bootstrap each par swap rate to back out the "
         "discount factor at that swap's maturity.\n"
         "Each new DF: prev_DF / (1 + rate × days/360). See the YieldCurve sheet.",
         f"YieldCurve!$F${data['_yc_last_pillar_row']}",
         f"DF at last pillar (≈{(data['leg3_end'].year - data['val_date'].year) + 30} years out)"),

        ("Step 3 — Forward 10Y CMS at every daily fixing",
         "The range accrual observes the 10Y CMS rate on every CALENDAR DAY in each "
         "accrual period. For each day d, the forward 10Y CMS rate is:\n"
         "      fwd(d) = (DF(d) − DF(d+10Y)) / annuity(d)\n"
         "where annuity(d) = Σ_{i=1..10} τ_i × DF(d+iY) and τ_i is the ACT/360 year "
         "fraction. See DailyFixings sheet, column Z.",
         "DailyFixings!Z:Z", "Forward 10Y CMS per day"),

        ("Step 4 — Calibrate SABR alpha",
         "SABR has 4 parameters: alpha (level), beta (skew, fixed at 40%), rho "
         "(correlation), nu (vol-of-vol). Rho, nu and beta are read from MarketData. "
         "Alpha is solved so that the SABR model reproduces the ATM swaption vol at "
         "the forward = par 10Y CMS rate.\n"
         "We start with α₀ = σ_ATM × f^(1-β) and iterate α_new = α × σ_ATM / SABR(α,…) "
         "twice. See DailyFixings columns AD–AF.",
         "DailyFixings!AF:AF", "Calibrated SABR α"),

        ("Step 5 — SABR vol at the barrier K = 4.2%",
         "Once α is calibrated, compute the SABR implied lognormal vol at strike K=4.2% "
         "using Hagan 2002 formula. This 'smile' vol is what the digital pricer uses. "
         "See DailyFixings column AG.",
         "DailyFixings!AG:AG", "σ_SABR at K = 4.2%"),

        ("Step 6 — Digital probability per day",
         "For each day d, the probability that the 10Y CMS sets at or below the 4.2% "
         "barrier is the Black digital floor:\n"
         "      P(S ≤ K) = N(-d₂),    d₂ = (ln(f/K) − σ²T/2) / (σ√T)\n"
         "Note: we use the PAR forward f (not the CMS convexity-adjusted forward) — this "
         "is the correct treatment for digital payoffs, verified to within 0.45% against "
         "the system on both valuation dates.",
         "DailyFixings!AI:AI", "Daily digital probability"),

        ("Step 7 — Aggregate to period coupons",
         "For each of 18 quarterly accrual periods:\n"
         "      Accrual % = average of daily probabilities in the period\n"
         "      Coupon PV = Notional × 5.32% × τ_30/360 × Accrual % × DF(pay date)\n"
         "where τ_30/360 = 0.25 (quarterly 30/360 day count). See PeriodSummary sheet.",
         f"PeriodSummary!$J${data['_ps_last_row']}",
         "Last period coupon PV (example; full schedule on PeriodSummary)"),

        ("Step 8 — Sum the 18 period PVs → Leg 3 PV",
         "Total Leg 3 PV = Σ period coupon PVs over the 18 quarterly periods from "
         "24-Oct-2026 to 24-Apr-2031.",
         data['_total_pv_cell'], "Leg 3 total PV"),
    ]

    for title, expl, cellref, note in steps:
        # Title row
        setcell(ws, r, 2, title, HEADER_FONT, fill=SECTION_FILL)
        setcell(ws, r, 3, "", HEADER_FONT, fill=SECTION_FILL)
        setcell(ws, r, 4, "", HEADER_FONT, fill=SECTION_FILL)
        r += 1
        # Explanation row
        setcell(ws, r, 2, expl, FORMULA_FONT, align=WRAP)
        if cellref:
            setcell(ws, r, 3, f"={cellref}", XREF_FONT, fmt='0.0000000000')
        if note:
            setcell(ws, r, 4, note, FORMULA_FONT, align=WRAP)
        # Make the row tall enough for the wrapped explanation
        ws.row_dimensions[r].height = max(15 * (1 + expl.count('\n')), 50)
        r += 2

    # Final result block
    r += 1
    setcell(ws, r, 2, "FINAL RESULT — Leg 3 PV", TITLE_FONT, fill=RESULT_FILL)
    setcell(ws, r, 3, f"={data['_total_pv_cell']}", HEADER_FONT, fmt='#,##0.00',
            fill=RESULT_FILL)
    setcell(ws, r, 4, "USD", FORMULA_FONT, fill=RESULT_FILL)
    r += 1
    setcell(ws, r, 2, "as % of notional", FORMULA_FONT)
    setcell(ws, r, 3, f"={data['_total_pv_cell']}/{data['_md_notional_cell']}",
            FORMULA_FONT, fmt='0.000%')

    r += 2
    setcell(ws, r, 2, f"Expected (per Python reference implementation): ${data['expected_pv']:,}",
            FORMULA_FONT, align=LEFT_ALIGN)


# =============================================================================
# SECTION 13 — DRIVER FUNCTION
# =============================================================================
# `build` ties everything together. Given a market data dict and an output
# path, it creates a fresh workbook and calls each sheet-builder in order.
# The order matters because each sheet's formulas reference cells created
# by earlier sheets (e.g. DailyDFs references YieldCurve pillars).

def build(data, out_path):
    """Build one Excel workbook and save it to out_path."""
    wb = Workbook()
    # Workbook() creates a fresh workbook with one default sheet. We remove it.
    wb.remove(wb.active)

    # The order here is important — each sheet's formulas may reference cells
    # set up by earlier sheets via the `data` dict.
    build_readme(wb, data)
    build_market_data(wb, data)     # populates _md_* cell references
    build_yield_curve(wb, data)     # uses _md_*; populates _yc_*
    build_daily_dfs(wb, data)       # uses _yc_*
    build_daily_fixings(wb, data)   # uses _md_* and DailyDFs sheet
    build_period_summary(wb, data)  # uses DailyFixings; populates _ps_* and _total_pv_cell
    build_valuation(wb, data)       # uses everything

    wb.save(out_path)
    print(f"Saved {out_path}")


# =============================================================================
# SECTION 14 — MAIN ENTRY POINT
# =============================================================================
# The `if __name__ == '__main__':` idiom means "only run this if THIS file is
# being executed directly" (not imported as a module). It's how Python scripts
# distinguish "run me as a program" from "import me as a library".

if __name__ == '__main__':
    import sys
    # `sys.argv` is the list of command-line arguments.
    # sys.argv[0] is always the script name (e.g. 'build_excel_explained.py').
    # sys.argv[1], if provided, is the first user argument.
    # If the user runs `python3 script.py apr21`, sys.argv = ['script.py', 'apr21'].
    arg = sys.argv[1] if len(sys.argv) > 1 else 'both'

    if arg == 'apr21':
        d = make_data_apr21()
        build(d, 'RangeAccrual_Apr21_2026.xlsx')
    elif arg == 'may7':
        d = make_data_may7()
        build(d, 'RangeAccrual_May7_2026.xlsx')
    elif arg == 'both':
        # Build both. Each call gets a FRESH data dict (important: don't share
        # because _md_* fields would get overwritten).
        d1 = make_data_apr21()
        build(d1, 'RangeAccrual_Apr21_2026.xlsx')
        d2 = make_data_may7()
        build(d2, 'RangeAccrual_May7_2026.xlsx')
    else:
        print(f"Unknown argument: {arg!r}. Use 'apr21', 'may7', or 'both'.")


# =============================================================================
# AFTER RUNNING
# =============================================================================
#
# 1. The script will save the .xlsx files in the current directory.
#
# 2. When you OPEN them in Excel or LibreOffice, the application will
#    automatically recalculate all formulas. You'll see the final PV in:
#
#       Valuation sheet → "FINAL RESULT — Leg 3 PV" row
#       PeriodSummary  → bottom of column J
#
# 3. To experiment: change any value on the MarketData sheet (blue cells)
#    and the entire workbook re-prices automatically. For example, try
#    changing the 4.2% barrier to 4.5% and watch the PV go up.
#
# 4. If you change anything programmatically with openpyxl (not Excel),
#    formulas are stored as TEXT and won't have values until next opened
#    in Excel/LibreOffice. There's a `recalc.py` script in the xlsx skill
#    that can force recalc via headless LibreOffice if needed.
#
# =============================================================================
